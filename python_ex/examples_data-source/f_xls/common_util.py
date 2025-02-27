"""
    현재 주식 매수 일지 시트를  읽어와서 정리해서 다시 Sheet로 정리한다. 
    loading from a file. 
    from openpyxl import load_workbook 
    wb = load_workbook( filename = 'empty_book.xlsx' )
    sheet_ranges = wb ['range names'] 

    먼저 특정 Sheet에서 ws 선택하기. 
    df = DataFrame(ws.values)

    XLS 데이터를 DB화 작업 한다. 
"""
from openpyxl import load_workbook 
import pandas as pd 
import pdb, os
from os import listdir 
from os.path import isfile, isdir, join 
import re
from bs4 import BeautifulSoup
import pandas as pd


dir_path = os.path.dirname(os.path.realpath(__file__))
data_krx_path = os.path.normpath( os.path.join( dir_path, 'data_krx' )) 
data_companyguide_path = os.path.normpath( os.path.join( dir_path, 'data_companyguide')) 
data_companyguide_path_1 = os.path.normpath( os.path.join( dir_path, 'data_companyguide', '20240320' )) 

def files_list( target_dir_path  ):
    onlyfiles = [ f for f in listdir( target_dir_path) if isfile( join( target_dir_path , f ))]
    print( onlyfiles ) 
    return onlyfiles 

def dirs_list( target_dir_path  ):
    onlydirs = [ f for f in listdir( target_dir_path) if isdir( join( target_dir_path , f ))]
    print( onlydirs ) 
    return onlydirs 

def load_f_xlsx( fileName , sheetName ):
    wb = load_workbook( filename = fileName ) 
    ws = wb[sheetName]
    df = pd.DataFrame( ws.values )
    df.columns = df.iloc[0] 
    df =  df.drop( df.index[0] )
    print(df) 
    return df 

def data_krx_last():
    f_list = files_list( data_krx_path ); 
    """
    for x in f_list:
        print( re.search(r'\d+.xlsx', x ).group())
    sorted_list = sorted( f_list , key=lambda x: int( re.search(r'\d+$', x ).group())) 
    """
    sorted_list = sorted( f_list , key=lambda x: x[-6::-13], reverse=True  ) 
    sorted_list.reverse() 
    print( sorted_list[-1] )
    last_file_path = os.path.normpath( os.path.join( data_krx_path , sorted_list[-1] ))
    df = load_f_xlsx( last_file_path,'Sheet1') 
    df = df.sort_values(by='배당수익률', ascending= False )
    return df 

def fnguide_last():
    last_file_path = os.path.normpath( os.path.join( dir_path , 'IdxRank_Excel.xls' ))
    df = load_f_xlsx( last_file_path,'IdxRank_Excel') 
    df = df.sort_values(by='부채비율', ascending= False )
    return df 
    

def fnguide_html_file():
    table = BeautifulSoup(open('./IdxRank_Excel.html','r').read()).find('table')
    df = pd.read_html(str(table)) #I think it accepts BeatifulSoup object
                             #otherwise try str(table) as input
    df = df[0].sort_values( by = '부채비율', ascending = True ) 
    return df                          

def data_companyguide_last_1():
    last_file_path = os.path.normpath( os.path.join( data_companyguide_path_1 , 'IdxRank_Excel.html' ))
    table = BeautifulSoup(open( last_file_path ,'r').read()).find('table')
    df = pd.read_html(str(table)) #I think it accepts BeatifulSoup object
                             #otherwise try str(table) as input
    df = df[0].sort_values( by = '부채비율', ascending = True ) 
    return df                          

def data_companyguide_last():
    dirs = dirs_list( data_companyguide_path ); 
    dirs = sorted( dirs, reverse = True )
    last_file_path = os.path.normpath( os.path.join( data_companyguide_path, dirs[0], 'IdxRank_Excel.xls' ))
    table = BeautifulSoup(open( last_file_path ,'r').read()).find('table')
    df = pd.read_html(str(table)) #I think it accepts BeatifulSoup object
                             #otherwise try str(table) as input
    df = df[0].sort_values( by = '부채비율', ascending = True ) 
    return df  

if __name__=='__main__':
    #df = data_krx_last();
    #df = fnguide_last(); 
    df = data_companyguide_last();
    print( df )




