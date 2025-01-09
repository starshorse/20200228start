"""
    현재 주식 매수 일지 시트를  읽어와서 정리해서 다시 Sheet로 정리한다. 
    loading from a file. 
    from openpyxl import load_workbook 
    wb = load_workbook( filename = 'empty_book.xlsx' )
    sheet_ranges = wb ['range names'] 

    먼저 특정 Sheet에서 ws 선택하기. 
    df = DataFrame(ws.values)

    XLS 데이터를 DB화 작업 한다. 
"""
from openpyxl import load_workbook 
import pandas as pd 
import toDB_stocks_01 as db_korsp 
import pdb

wb = load_workbook( filename = 'stocks.xlsx') 
#ws = wb['매매일지']
#ws = wb['유가증권_상장사목록']
ws = wb['코스닥_상장사목록']

df = pd.DataFrame( ws.values )
df.columns = df.iloc[0] 
df =  df.drop( df.index[0] )
df['종목코드'] = df['종목코드'].astype(str).str.pad(6 , fillchar='0' )
df['주요제품'] = df['주요제품'].astype(str).str.replace("'","_")

db_korsp.merge_stocklist( df, '[dbo].[코스닥_상장사목록]' )
pdb.set_trace() 

print(df) 

