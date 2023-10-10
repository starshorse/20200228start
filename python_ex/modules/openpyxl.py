from openpyxl import load_workbook , Workbook 

def getWork_book( xlsx_name ):
    return load_workbook( xlsx_name ) 

def getSheet( workbook, sheetName ):
    return  workbook[sheetName]; 

def getMax_row( sheet ):
    return sheet.max_row 

def append( sheet, data ):
    return sheet.append( data ); 

def save( sheet, fileName ):
    return sheet.save( fileName );   

