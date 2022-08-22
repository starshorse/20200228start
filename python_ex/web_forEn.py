from selenium import webdriver
from selenium.webdriver.common.by import By
from os import listdir 
from openpyxl import load_workbook , Workbook 
from time import sleep 

driver = webdriver.Chrome('./chromedriver')
driver.get('https://finance.daum.net/domestic/influential_investors')
sleep(3) 

items = driver.find_element('class name','box_contents')
#print( items.text )
def list_chuck(arr, n):
    return [arr[i: i + n] for i in range(0, len(arr), n)]

items = driver.find_element('xpath','//*[@id="boxInfluentialInvestors"]/div[2]/div[1]/table')
# print( items.text )

list_items_info = items.text.splitlines()
# print( list_items_info )

del list_items_info[:2]
# print( list_items_info )

forEn = list_chuck( list_items_info, 4 )

print( forEn )
this_date = driver.find_element('xpath', '//*[@id="labDescription"]/em' )
print( this_date.text )
this_date_1 = this_date.text[6:]
this_date_mmdd = this_date_1.split('.')
print( this_date_mmdd )
from datetime import date 
this_year = date.today().year 
print( this_year )
this_date_2 = str(this_year)+"-"+this_date_mmdd[0]+"-"+this_date_mmdd[1]
print( this_date_2 )
for idx, x in enumerate( forEn ):
    if idx%2:
        x.append(-1)
    else :
        x.append(1)
    x.append( this_date_2 )
    
print( forEn )
result_xlsx_add = load_workbook('forEn.xlsx')
result_sheet_add = result_xlsx_add['Sheet']

max = result_sheet_add.max_row 
print( max )

for x in forEn:
    result_sheet_add.append(x)
    
result_xlsx_add.save('forEn.xlsx')